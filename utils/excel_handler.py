"""
Excel Template Generator & Parser for Report Card marks upload.
Generates pre-filled Excel templates and parses uploaded mark sheets.
"""
import io
import uuid
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Grade calculation (same as existing system)
def calculate_grade(percentage):
    if percentage >= 91: return "A1"
    elif percentage >= 81: return "A2"
    elif percentage >= 71: return "B1"
    elif percentage >= 61: return "B2"
    elif percentage >= 51: return "C1"
    elif percentage >= 41: return "C2"
    elif percentage >= 33: return "D"
    else: return "E"


def generate_marks_template(students, subject_name, exam_name, class_name, max_marks=100):
    """
    Generate an Excel template pre-filled with student names for marks entry.
    Returns bytes of the Excel file.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Marks Entry"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    subheader_font = Font(name="Calibri", bold=True, size=10)
    subheader_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    locked_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")

    # Title row
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{exam_name} — {subject_name} — {class_name}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = center

    # Info row
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Max Marks: {max_marks} | Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=9, italic=True)
    ws["A2"].alignment = center

    # Instruction row
    ws.merge_cells("A3:F3")
    ws["A3"] = "Fill marks in column D. Use AB=Absent, ML=Medical Leave, NA=Not Applicable, EX=Exempted in column E."
    ws["A3"].font = Font(name="Calibri", size=9, color="EF4444")
    ws["A3"].alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # Column headers (row 4)
    headers = ["S.No.", "Student ID", "Student Name", "Marks", "Special Code", "Remarks"]
    col_widths = [8, 38, 30, 12, 15, 25]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Student data rows
    for row_idx, student in enumerate(students, 5):
        # S.No.
        cell = ws.cell(row=row_idx, column=1, value=row_idx - 4)
        cell.font = data_font
        cell.alignment = center
        cell.fill = locked_fill
        cell.border = thin_border

        # Student ID (hidden UUID)
        cell = ws.cell(row=row_idx, column=2, value=str(student["id"]))
        cell.font = Font(name="Calibri", size=8, color="94A3B8")
        cell.fill = locked_fill
        cell.border = thin_border

        # Student Name
        name = f"{student.get('first_name', '')} {student.get('last_name', '')}".strip()
        if student.get("roll_number"):
            name = f"[{student['roll_number']}] {name}"
        cell = ws.cell(row=row_idx, column=3, value=name)
        cell.font = data_font
        cell.fill = locked_fill
        cell.border = thin_border

        # Marks (editable)
        cell = ws.cell(row=row_idx, column=4, value="")
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = center
        cell.border = thin_border

        # Special Code (editable)
        cell = ws.cell(row=row_idx, column=5, value="")
        cell.font = data_font
        cell.alignment = center
        cell.border = thin_border

        # Remarks (editable)
        cell = ws.cell(row=row_idx, column=6, value="")
        cell.font = data_font
        cell.border = thin_border

    # Freeze panes
    ws.freeze_panes = "A5"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def parse_marks_upload(file_bytes, max_marks=100):
    """
    Parse an uploaded Excel file with marks.
    Returns list of dicts: [{student_id, marks, special_code, remarks}, ...]
    Also returns errors list if any validation fails.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required")

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    records = []
    errors = []
    row_num = 0

    for row in ws.iter_rows(min_row=5, values_only=False):
        row_num += 1
        # Get cell values
        cells = [cell.value for cell in row[:6]]
        sno, student_id, name, marks_val, special_code, remarks = (
            cells[0] if len(cells) > 0 else None,
            cells[1] if len(cells) > 1 else None,
            cells[2] if len(cells) > 2 else None,
            cells[3] if len(cells) > 3 else None,
            cells[4] if len(cells) > 4 else None,
            cells[5] if len(cells) > 5 else None,
        )

        # Skip empty rows
        if not student_id:
            continue

        # Validate student ID is UUID
        try:
            sid = str(student_id).strip()
            uuid.UUID(sid)
        except (ValueError, AttributeError):
            errors.append(f"Row {row_num + 4}: Invalid student ID '{student_id}'")
            continue

        record = {
            "student_id": sid,
            "marks_obtained": None,
            "special_code": None,
            "remarks": str(remarks).strip() if remarks else None,
        }

        # Check special code
        if special_code:
            code = str(special_code).strip().upper()
            if code in ("AB", "ML", "NA", "EX"):
                record["special_code"] = code
                records.append(record)
                continue
            else:
                errors.append(f"Row {row_num + 4}: Invalid special code '{special_code}'. Use AB/ML/NA/EX.")
                continue

        # Parse marks
        if marks_val is not None and str(marks_val).strip() != "":
            try:
                marks = float(marks_val)
                if marks < 0:
                    errors.append(f"Row {row_num + 4}: Marks cannot be negative ({marks})")
                    continue
                if marks > max_marks:
                    errors.append(f"Row {row_num + 4}: Marks ({marks}) exceed max marks ({max_marks})")
                    continue
                record["marks_obtained"] = marks
            except (ValueError, TypeError):
                # Check if it's a special code in marks column
                code = str(marks_val).strip().upper()
                if code in ("AB", "ML", "NA", "EX"):
                    record["special_code"] = code
                else:
                    errors.append(f"Row {row_num + 4}: Invalid marks value '{marks_val}'")
                    continue

        records.append(record)

    wb.close()
    return records, errors
