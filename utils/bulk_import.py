"""
VedaFlow Bulk Import — Excel template generators, parsers, and database inserter.
Used by the Easy Setup wizard to create a full school from Excel uploads.

Dependency order: Organization -> Branch -> AcademicYear -> Class -> Section ->
                   Subject -> Users -> Teachers -> Students -> Transport
"""

import io
import re
import uuid
import logging
from datetime import datetime, date

logger = logging.getLogger("bulk_import")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Styles (matching excel_handler.py) ──────────────────────
HEADER_FONT = None
HEADER_FILL = None
SUBHEADER_FONT = None
SUBHEADER_FILL = None
DATA_FONT = None
THIN_BORDER = None
CENTER = None
REQUIRED_FILL = None

def _init_styles():
    global HEADER_FONT, HEADER_FILL, SUBHEADER_FONT, SUBHEADER_FILL, DATA_FONT, THIN_BORDER, CENTER, REQUIRED_FILL
    HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10)
    SUBHEADER_FILL = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    DATA_FONT = Font(name="Calibri", size=10)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    CENTER = Alignment(horizontal="center", vertical="center")
    REQUIRED_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

def _check_openpyxl():
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    _init_styles()


# ═══════════════════════════════════════════════════════════
# TEMPLATE GENERATORS — Download Excel templates
# ═══════════════════════════════════════════════════════════

def generate_admin_template() -> bytes:
    """Generate Excel template for Admin/Principal upload."""
    _check_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Admin Staff"

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Admin / Principal Details — Upload Template"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER

    # Instructions
    ws.merge_cells("A2:D2")
    ws["A2"] = "Fields with * are required. Role must be: Principal or Admin."
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="EF4444")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[2].height = 25

    # Headers
    headers = ["Name *", "Email *", "Phone *", "Role *"]
    widths = [30, 35, 18, 15]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sample row
    samples = ["Dr. Rajesh Kumar", "rajesh@school.edu.in", "9876543210", "Principal"]
    for i, val in enumerate(samples, 1):
        cell = ws.cell(row=4, column=i, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="9CA3AF")
        cell.border = THIN_BORDER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_teacher_template() -> bytes:
    """Generate Excel template for Teacher upload."""
    _check_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"

    ws.merge_cells("A1:H1")
    ws["A1"] = "Teacher Details — Upload Template"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:H2")
    ws["A2"] = "Fields with * are required. Phone must be 10 digits. Email must be unique."
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="EF4444")
    ws.row_dimensions[2].height = 25

    headers = ["First Name *", "Last Name", "Email *", "Phone *", "Subject *",
               "Qualification", "Experience (Years)", "Monthly Salary"]
    widths = [20, 20, 30, 18, 20, 25, 15, 15]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(i)].width = w

    samples = ["Priya", "Sharma", "priya@school.edu.in", "9876543211", "Mathematics",
               "M.Sc., B.Ed.", "8", "45000"]
    for i, val in enumerate(samples, 1):
        cell = ws.cell(row=4, column=i, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="9CA3AF")
        cell.border = THIN_BORDER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_student_template() -> bytes:
    """Generate Excel template for Student upload."""
    _check_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    ws.merge_cells("A1:P1")
    ws["A1"] = "Student Details — Upload Template"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:P2")
    ws["A2"] = "Fields with * are required. DOB format: DD/MM/YYYY. Gender: Male/Female/Other. Father phone must be 10 digits."
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="EF4444")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = [
        "Class *", "Section *", "Roll No *", "First Name *", "Last Name *",
        "Gender *", "DOB (DD/MM/YYYY)", "Admission No", "Father Name *",
        "Father Phone *", "Father Email", "Mother Name", "Mother Phone",
        "Address", "City", "State",
    ]
    widths = [10, 10, 10, 18, 18, 10, 16, 15, 20, 16, 25, 20, 16, 30, 15, 15]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(i)].width = w

    samples = [
        "10", "A", "001", "Aarav", "Patel", "Male", "15/03/2012", "ADM2024001",
        "Rajesh Patel", "9876543212", "rajesh@gmail.com", "Sunita Patel",
        "9876543213", "123 MG Road", "Mumbai", "Maharashtra",
    ]
    for i, val in enumerate(samples, 1):
        cell = ws.cell(row=4, column=i, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="9CA3AF")
        cell.border = THIN_BORDER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_transport_template() -> bytes:
    """Generate Excel template for Transport upload."""
    _check_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Transport"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Transport Details — Upload Template"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:E2")
    ws["A2"] = "Fields with * are required. Stops: semicolon-separated list (at least 2 stops per route)."
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="EF4444")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[2].height = 25

    headers = ["Vehicle Number *", "Driver Name *", "Driver Phone *", "Route Name *", "Stops (semicolon-separated) *"]
    widths = [18, 22, 18, 22, 50]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(i)].width = w

    samples = ["MH12AB1234", "Ramesh Yadav", "9876543214", "Route 1 - North",
               "School Gate; MG Road; Railway Station; City Center; Bus Stand"]
    for i, val in enumerate(samples, 1):
        cell = ws.cell(row=4, column=i, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True, color="9CA3AF")
        cell.border = THIN_BORDER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════
# PARSERS — Validate Excel uploads
# ═══════════════════════════════════════════════════════════

def _clean_str(val):
    """Clean cell value to string."""
    if val is None:
        return ""
    return str(val).strip()

def _clean_phone(val):
    """Clean phone number — remove +91, spaces, dashes."""
    phone = _clean_str(val)
    phone = re.sub(r"[^0-9]", "", phone)
    if phone.startswith("91") and len(phone) == 12:
        phone = phone[2:]
    return phone

def _validate_email(email):
    """Basic email validation."""
    if not email:
        return False
    return bool(re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email))


def parse_admin_excel(file_bytes: bytes) -> dict:
    """Parse admin/principal Excel upload. Returns {valid, data, errors}."""
    _check_openpyxl()
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    data = []
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not row or all(v is None for v in row):
            continue

        name = _clean_str(row[0] if len(row) > 0 else "")
        email = _clean_str(row[1] if len(row) > 1 else "").lower()
        phone = _clean_phone(row[2] if len(row) > 2 else "")
        role = _clean_str(row[3] if len(row) > 3 else "").title()

        row_errors = []
        if not name:
            row_errors.append("Name is required")
        if not email:
            row_errors.append("Email is required")
        elif not _validate_email(email):
            row_errors.append(f"Invalid email: {email}")
        if not phone:
            row_errors.append("Phone is required")
        elif len(phone) != 10:
            row_errors.append(f"Phone must be 10 digits: {phone}")
        if role not in ("Principal", "Admin"):
            row_errors.append(f"Role must be Principal or Admin: {role}")

        if row_errors:
            errors.append({"row": row_idx, "errors": row_errors})
        data.append({"name": name, "email": email, "phone": phone, "role": role, "row": row_idx})

    wb.close()
    return {"valid": len(errors) == 0 and len(data) > 0, "data": data, "errors": errors, "count": len(data)}


def parse_teacher_excel(file_bytes: bytes) -> dict:
    """Parse teacher Excel upload."""
    _check_openpyxl()
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    data = []
    errors = []
    seen_emails = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not row or all(v is None for v in row):
            continue

        first_name = _clean_str(row[0] if len(row) > 0 else "")
        last_name = _clean_str(row[1] if len(row) > 1 else "")
        email = _clean_str(row[2] if len(row) > 2 else "").lower()
        phone = _clean_phone(row[3] if len(row) > 3 else "")
        subject = _clean_str(row[4] if len(row) > 4 else "")
        qualification = _clean_str(row[5] if len(row) > 5 else "")
        experience = _clean_str(row[6] if len(row) > 6 else "")
        salary = _clean_str(row[7] if len(row) > 7 else "")

        row_errors = []
        if not first_name:
            row_errors.append("First Name is required")
        if not email:
            row_errors.append("Email is required")
        elif not _validate_email(email):
            row_errors.append(f"Invalid email: {email}")
        elif email in seen_emails:
            row_errors.append(f"Duplicate email: {email}")
        if not phone:
            row_errors.append("Phone is required")
        elif len(phone) != 10:
            row_errors.append(f"Phone must be 10 digits: {phone}")
        if not subject:
            row_errors.append("Subject is required")

        seen_emails.add(email)

        if row_errors:
            errors.append({"row": row_idx, "errors": row_errors})
        data.append({
            "first_name": first_name, "last_name": last_name, "email": email,
            "phone": phone, "subject": subject, "qualification": qualification,
            "experience": experience, "salary": salary, "row": row_idx,
        })

    wb.close()
    return {"valid": len(errors) == 0 and len(data) > 0, "data": data, "errors": errors, "count": len(data)}


def parse_student_excel(file_bytes: bytes) -> dict:
    """Parse student Excel upload."""
    _check_openpyxl()
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    data = []
    errors = []
    seen_admission = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not row or all(v is None for v in row):
            continue

        class_name = _clean_str(row[0] if len(row) > 0 else "")
        section = _clean_str(row[1] if len(row) > 1 else "").upper()
        roll_no = _clean_str(row[2] if len(row) > 2 else "")
        first_name = _clean_str(row[3] if len(row) > 3 else "")
        last_name = _clean_str(row[4] if len(row) > 4 else "")
        gender = _clean_str(row[5] if len(row) > 5 else "").title()
        dob_str = _clean_str(row[6] if len(row) > 6 else "")
        admission_no = _clean_str(row[7] if len(row) > 7 else "")
        father_name = _clean_str(row[8] if len(row) > 8 else "")
        father_phone = _clean_phone(row[9] if len(row) > 9 else "")
        father_email = _clean_str(row[10] if len(row) > 10 else "").lower()
        mother_name = _clean_str(row[11] if len(row) > 11 else "")
        mother_phone = _clean_phone(row[12] if len(row) > 12 else "")
        address = _clean_str(row[13] if len(row) > 13 else "")
        city = _clean_str(row[14] if len(row) > 14 else "")
        state = _clean_str(row[15] if len(row) > 15 else "")

        row_errors = []
        if not class_name:
            row_errors.append("Class is required")
        if not section:
            row_errors.append("Section is required")
        if not roll_no:
            row_errors.append("Roll No is required")
        if not first_name:
            row_errors.append("First Name is required")
        if not last_name:
            row_errors.append("Last Name is required")
        if gender not in ("Male", "Female", "Other"):
            row_errors.append(f"Gender must be Male/Female/Other: {gender}")
        if not father_name:
            row_errors.append("Father Name is required")
        if not father_phone:
            row_errors.append("Father Phone is required")
        elif len(father_phone) != 10:
            row_errors.append(f"Father Phone must be 10 digits: {father_phone}")

        # Parse DOB
        dob = None
        if dob_str:
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y"):
                try:
                    dob = datetime.strptime(dob_str, fmt).date()
                    break
                except ValueError:
                    continue
            if not dob:
                row_errors.append(f"Invalid DOB format: {dob_str} (use DD/MM/YYYY)")

        # Check admission_no uniqueness within file
        if admission_no:
            if admission_no in seen_admission:
                row_errors.append(f"Duplicate Admission No: {admission_no}")
            seen_admission.add(admission_no)

        if row_errors:
            errors.append({"row": row_idx, "errors": row_errors})

        data.append({
            "class_name": class_name, "section": section, "roll_no": roll_no,
            "first_name": first_name, "last_name": last_name, "gender": gender,
            "dob": dob.isoformat() if dob else None, "admission_no": admission_no,
            "father_name": father_name, "father_phone": father_phone,
            "father_email": father_email, "mother_name": mother_name,
            "mother_phone": mother_phone, "address": address, "city": city,
            "state": state, "row": row_idx,
        })

    wb.close()
    return {"valid": len(errors) == 0 and len(data) > 0, "data": data, "errors": errors, "count": len(data)}


def parse_transport_excel(file_bytes: bytes) -> dict:
    """Parse transport Excel upload."""
    _check_openpyxl()
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    data = []
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not row or all(v is None for v in row):
            continue

        vehicle_no = _clean_str(row[0] if len(row) > 0 else "").upper()
        driver_name = _clean_str(row[1] if len(row) > 1 else "")
        driver_phone = _clean_phone(row[2] if len(row) > 2 else "")
        route_name = _clean_str(row[3] if len(row) > 3 else "")
        stops_str = _clean_str(row[4] if len(row) > 4 else "")

        stops = [s.strip() for s in stops_str.split(";") if s.strip()] if stops_str else []

        row_errors = []
        if not vehicle_no:
            row_errors.append("Vehicle Number is required")
        if not driver_name:
            row_errors.append("Driver Name is required")
        if not driver_phone:
            row_errors.append("Driver Phone is required")
        elif len(driver_phone) != 10:
            row_errors.append(f"Driver Phone must be 10 digits: {driver_phone}")
        if not route_name:
            row_errors.append("Route Name is required")
        if len(stops) < 2:
            row_errors.append("At least 2 stops are required (semicolon-separated)")

        if row_errors:
            errors.append({"row": row_idx, "errors": row_errors})

        data.append({
            "vehicle_no": vehicle_no, "driver_name": driver_name,
            "driver_phone": driver_phone, "route_name": route_name,
            "stops": stops, "row": row_idx,
        })

    wb.close()
    return {"valid": len(errors) == 0 and len(data) > 0, "data": data, "errors": errors, "count": len(data)}


# ═══════════════════════════════════════════════════════════
# BULK INSERTER — Create all records in one transaction
# ═══════════════════════════════════════════════════════════

def _generate_password(length=12):
    """Generate a random strong password."""
    import random
    import string
    chars = string.ascii_letters + string.digits + "!@#$%"
    # Ensure at least one of each type
    pwd = [
        random.choice(string.ascii_uppercase),
        random.choice(string.ascii_lowercase),
        random.choice(string.digits),
        random.choice("!@#$%"),
    ]
    pwd += [random.choice(chars) for _ in range(length - 4)]
    random.shuffle(pwd)
    return "".join(pwd)


async def execute_school_setup(db, setup_data: dict) -> dict:
    """
    Create a full school (org + branch + users + students + transport) in one transaction.

    setup_data: {
        org: {name, email, phone, address, city, state, pincode},
        branch: {name, board_type, city, state, principal_name},
        admins: [{name, email, phone, role}, ...],
        teachers: [{first_name, last_name, email, phone, subject, ...}, ...],
        students: [{class_name, section, roll_no, first_name, last_name, ...}, ...],
        transport: [{vehicle_no, driver_name, driver_phone, route_name, stops}, ...],  # optional
    }

    Returns: {success: bool, credentials: [...], org_id, branch_id, error?}
    """
    from models.organization import Organization
    from models.branch import Branch
    from models.academic import AcademicYear, Class, Section, Subject
    from models.user import User, UserRole
    from models.teacher import Teacher
    from models.student import Student
    from models.transport import Vehicle, TransportRoute, RouteStop
    from utils.auth import hash_password
    import re

    credentials = []
    org_data = setup_data.get("org", {})
    branch_data = setup_data.get("branch", {})
    admin_list = setup_data.get("admins", [])
    teacher_list = setup_data.get("teachers", [])
    student_list = setup_data.get("students", [])
    transport_list = setup_data.get("transport", [])

    try:
        # 1. Create Organization
        org_name = org_data.get("name", "New School")
        slug = re.sub(r"[^a-z0-9]+", "-", org_name.lower()).strip("-")
        # Ensure unique slug
        from sqlalchemy import select as sa_select
        existing = await db.scalar(sa_select(Organization).where(Organization.slug == slug))
        if existing:
            slug = f"{slug}-{uuid.uuid4().hex[:6]}"

        org = Organization(
            name=org_name,
            slug=slug,
            email=org_data.get("email", ""),
            phone=org_data.get("phone", ""),
            address=org_data.get("address", ""),
            city=org_data.get("city", ""),
            state=org_data.get("state", ""),
            pincode=org_data.get("pincode", ""),
            is_active=True,
        )
        db.add(org)
        await db.flush()
        org_id = org.id

        # 2. Create Branch
        branch = Branch(
            org_id=org_id,
            name=branch_data.get("name", org_name),
            code=slug[:10].upper(),
            email=org_data.get("email", ""),
            phone=org_data.get("phone", ""),
            city=branch_data.get("city", org_data.get("city", "")),
            state=branch_data.get("state", org_data.get("state", "")),
            principal_name=branch_data.get("principal_name", ""),
            is_active=True,
        )
        # Set board_type if provided
        board = branch_data.get("board_type", "")
        if board:
            try:
                from models.branch import BoardType
                branch.board_type = BoardType(board.lower())
            except (ValueError, KeyError):
                pass
        db.add(branch)
        await db.flush()
        branch_id = branch.id

        # 3. Create Academic Year
        current_year = datetime.now().year
        ay_name = f"{current_year}-{str(current_year + 1)[-2:]}"
        academic_year = AcademicYear(
            branch_id=branch_id,
            name=ay_name,
            start_date=date(current_year, 4, 1),
            end_date=date(current_year + 1, 3, 31),
            is_current=True,
            is_active=True,
        )
        db.add(academic_year)
        await db.flush()
        ay_id = academic_year.id

        # 4. Create Classes and Sections from student data
        class_map = {}   # class_name -> Class object
        section_map = {} # "class_name|section" -> Section object

        unique_classes = set()
        unique_sections = set()
        for s in student_list:
            cn = s.get("class_name", "")
            sec = s.get("section", "")
            if cn:
                unique_classes.add(cn)
            if cn and sec:
                unique_sections.add(f"{cn}|{sec}")

        for i, cn in enumerate(sorted(unique_classes)):
            cls = Class(
                branch_id=branch_id,
                name=cn,
                display_order=i + 1,
                is_active=True,
            )
            db.add(cls)
            await db.flush()
            class_map[cn] = cls

        for sec_key in sorted(unique_sections):
            cn, sec_name = sec_key.split("|", 1)
            cls_obj = class_map.get(cn)
            if cls_obj:
                section = Section(
                    class_id=cls_obj.id,
                    branch_id=branch_id,
                    name=sec_name,
                    is_active=True,
                )
                db.add(section)
                await db.flush()
                section_map[sec_key] = section

        # 5. Create Subjects from teacher data
        subject_map = {}  # subject_name -> Subject object
        unique_subjects = set()
        for t in teacher_list:
            subj = t.get("subject", "")
            if subj:
                unique_subjects.add(subj)

        for subj_name in sorted(unique_subjects):
            subject = Subject(
                branch_id=branch_id,
                name=subj_name,
                code=subj_name[:5].upper(),
                is_active=True,
            )
            db.add(subject)
            await db.flush()
            subject_map[subj_name] = subject

        # 6. Create Admin Users
        for admin in admin_list:
            password = _generate_password()
            name_parts = admin.get("name", "Admin").split(" ", 1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ""

            user = User(
                org_id=org_id,
                branch_id=branch_id,
                email=admin.get("email", ""),
                phone=admin.get("phone", ""),
                first_name=first_name,
                last_name=last_name,
                password_hash=hash_password(password),
                role=UserRole.SCHOOL_ADMIN,
                is_active=True,
                is_verified=True,
            )
            db.add(user)
            await db.flush()

            credentials.append({
                "name": admin.get("name", ""),
                "email": admin.get("email", ""),
                "phone": admin.get("phone", ""),
                "role": admin.get("role", "Admin"),
                "password": password,
                "user_id": str(user.id),
            })

        # 7. Create Teacher Users + Teacher records
        for t in teacher_list:
            password = _generate_password()

            user = User(
                org_id=org_id,
                branch_id=branch_id,
                email=t.get("email", ""),
                phone=t.get("phone", ""),
                first_name=t.get("first_name", ""),
                last_name=t.get("last_name", ""),
                password_hash=hash_password(password),
                role=UserRole.TEACHER,
                is_active=True,
                is_verified=True,
            )
            db.add(user)
            await db.flush()

            teacher = Teacher(
                user_id=user.id,
                branch_id=branch_id,
                employee_id=f"TCH{str(uuid.uuid4().hex[:6]).upper()}",
                qualification=t.get("qualification", ""),
                is_active=True,
            )
            # Link subject if exists
            subj = subject_map.get(t.get("subject", ""))
            if subj:
                teacher.subject_id = subj.id
            db.add(teacher)
            await db.flush()

            credentials.append({
                "name": f"{t.get('first_name', '')} {t.get('last_name', '')}".strip(),
                "email": t.get("email", ""),
                "phone": t.get("phone", ""),
                "role": "Teacher",
                "password": password,
                "user_id": str(user.id),
            })

        # 8. Create Student Users + Student records
        for s in student_list:
            password = _generate_password()
            # Generate email if not provided — use phone or admission_no based
            student_email = s.get("father_email", "")
            if not student_email:
                student_email = f"{s.get('first_name', 'student').lower()}.{s.get('roll_no', '')}@{slug}.school"

            user = User(
                org_id=org_id,
                branch_id=branch_id,
                email=student_email,
                phone=s.get("father_phone", ""),
                first_name=s.get("first_name", ""),
                last_name=s.get("last_name", ""),
                password_hash=hash_password(password),
                role=UserRole.STUDENT,
                is_active=True,
                is_verified=True,
            )
            db.add(user)
            await db.flush()

            # Get class and section IDs
            cls_obj = class_map.get(s.get("class_name", ""))
            sec_key = f"{s.get('class_name', '')}|{s.get('section', '')}"
            sec_obj = section_map.get(sec_key)

            student = Student(
                user_id=user.id,
                branch_id=branch_id,
                class_id=cls_obj.id if cls_obj else None,
                section_id=sec_obj.id if sec_obj else None,
                academic_year_id=ay_id,
                roll_number=s.get("roll_no", ""),
                admission_number=s.get("admission_no", ""),
                father_name=s.get("father_name", ""),
                father_phone=s.get("father_phone", ""),
                father_email=s.get("father_email", ""),
                mother_name=s.get("mother_name", ""),
                mother_phone=s.get("mother_phone", ""),
                address=s.get("address", ""),
                city=s.get("city", ""),
                state=s.get("state", ""),
                is_active=True,
            )
            # Set gender
            gender = s.get("gender", "")
            if gender:
                try:
                    from models.student import Gender
                    student.gender = Gender(gender.upper())
                except (ValueError, KeyError):
                    pass
            # Set DOB
            if s.get("dob"):
                try:
                    student.date_of_birth = date.fromisoformat(s["dob"])
                except (ValueError, TypeError):
                    pass

            db.add(student)
            await db.flush()

            credentials.append({
                "name": f"{s.get('first_name', '')} {s.get('last_name', '')}".strip(),
                "email": student_email,
                "phone": s.get("father_phone", ""),
                "role": "Student",
                "password": password,
                "user_id": str(user.id),
                "class": s.get("class_name", ""),
                "section": s.get("section", ""),
                "roll_no": s.get("roll_no", ""),
            })

        # 9. Create Transport (if provided)
        for tr in transport_list:
            vehicle = Vehicle(
                branch_id=branch_id,
                vehicle_number=tr.get("vehicle_no", ""),
                driver_name=tr.get("driver_name", ""),
                driver_phone=tr.get("driver_phone", ""),
                is_active=True,
            )
            db.add(vehicle)
            await db.flush()

            route = TransportRoute(
                branch_id=branch_id,
                vehicle_id=vehicle.id,
                name=tr.get("route_name", ""),
                is_active=True,
            )
            db.add(route)
            await db.flush()

            for i, stop_name in enumerate(tr.get("stops", [])):
                stop = RouteStop(
                    route_id=route.id,
                    name=stop_name,
                    stop_order=i + 1,
                )
                db.add(stop)

        await db.commit()

        logger.info(
            f"[EASY-SETUP] School created: {org_name} — "
            f"{len(admin_list)} admins, {len(teacher_list)} teachers, "
            f"{len(student_list)} students, {len(transport_list)} transport routes"
        )

        return {
            "success": True,
            "org_id": str(org_id),
            "branch_id": str(branch_id),
            "credentials": credentials,
            "summary": {
                "admins": len(admin_list),
                "teachers": len(teacher_list),
                "students": len(student_list),
                "classes": len(class_map),
                "sections": len(section_map),
                "subjects": len(subject_map),
                "transport_routes": len(transport_list),
            },
        }

    except Exception as e:
        await db.rollback()
        logger.error(f"[EASY-SETUP] Failed: {e}")
        return {"success": False, "error": str(e)}


async def mass_send_credentials(db, branch_id, credentials_list: list) -> dict:
    """Send login credentials via email to all users."""
    from utils.notifier import send_platform_email
    import uuid as uuid_mod

    sent = 0
    failed = 0
    bid = str(branch_id)

    for cred in credentials_list:
        email = cred.get("email", "")
        if not email or "@" not in email:
            failed += 1
            continue

        name = cred.get("name", "User")
        role = cred.get("role", "User")
        password = cred.get("password", "")

        subject = "Your VedaSchoolPro Login Credentials"
        body = f"""
Dear {name},

Your school has been registered on VedaSchoolPro — School Management System.

Your login credentials:
  Email: {email}
  Password: {password}
  Role: {role}

Login URL: https://app.vedaschoolpro.in/login

Please change your password after first login.

Regards,
VedaSchoolPro Team
"""
        try:
            await send_platform_email(
                db=db,
                branch_id=bid,
                to_email=email,
                subject=subject,
                body=body,
            )
            sent += 1
        except Exception as e:
            logger.warning(f"[EASY-SETUP] Email failed for {email}: {e}")
            failed += 1

    return {"sent": sent, "failed": failed, "total": len(credentials_list)}
